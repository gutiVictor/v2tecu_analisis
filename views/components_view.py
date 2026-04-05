import streamlit as st
import pandas as pd
from typing import List, Dict
import io
from views.constants import COLOR_CUMPLE, COLOR_NO_CUMPLE, COLOR_PTE, COLOR_PRIMARY, UMBRALES_ALERTAS
import logging
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

def color_tipo(tipo: str) -> str:
    """
    Mapea tipo de mensaje a color hexadecimal para UI consistente.
    
    Args:
        tipo: String con categoría ('success', 'warning', 'error', 'info')
        
    Returns:
        String con código de color hexadecimal
    """
    return {
        'success': COLOR_CUMPLE,
        'warning': COLOR_PTE,
        'error': COLOR_NO_CUMPLE,
        'info': '#38bdf8',
    }.get(tipo, COLOR_PRIMARY)  # Fallback al color primario si no coincide


def _preparar_datos_para_click(df_filtrado: pd.DataFrame, columnas_clave: List[str]) -> pd.DataFrame:
    """
    Crea un identificador único por fila para rastrear clicks en gráficos.
    """
    df_click = df_filtrado.copy()  # Evitar modificar el original
    # Concatenar valores de columnas clave como string único por fila
    df_click['_click_id'] = df_click[columnas_clave].astype(str).agg('_'.join, axis=1)
    return df_click

def mostrar_datos_fuente(
    df_filtrado: pd.DataFrame, 
    seleccion: Dict, 
    columnas_filtro: List[tuple], 
    titulo_seccion: str = "🔍 Datos Fuente"
) -> None:
    """
    Muestra en un expandable los registros que generaron el elemento clickeado.
    """
    # Validar que haya una selección válida con puntos
    if not seleccion or 'points' not in seleccion or not seleccion['points']:
        return
    
    punto = seleccion['points'][0]  # Tomar el primer punto seleccionado
    
    # Solo procesar si el punto tiene customdata (metadatos del gráfico)
    if 'customdata' in punto:
        with st.expander(titulo_seccion, expanded=True):
            df_resultado = df_filtrado.copy()  # Copia para no alterar original
            
            # Aplicar filtros dinámicos según los valores del punto clickeado
            for i, (col, _) in enumerate(columnas_filtro):
                if i < len(punto.get('customdata', [])):
                    valor = punto['customdata'][i]
                    if pd.notna(valor):  # Ignorar valores nulos
                        # Filtrar por coincidencia exacta de string (robusto a tipos mixtos)
                        df_resultado = df_resultado[
                            df_resultado[col].astype(str) == str(valor)
                        ]
            
            # Mostrar contador de registros encontrados
            st.caption(f"Registros que generan este punto: {len(df_resultado)}")
            
            # Seleccionar columnas relevantes para mostrar al usuario
            cols_display = [
                c for c in ['No_Orden', 'Cliente', 'Ciudad', 'Transportadora', 
                           'Fecha_Entrega', 'Cumple_NNS', 'Desvio_Entrega', 'Area_Incumple'] 
                if c in df_resultado.columns
            ]
            
            # Tabla interactiva con los 50 primeros registros
            st.dataframe(
                df_resultado[cols_display].head(50), 
                use_container_width=True, 
                hide_index=True
            )
            
            # Botón de exportación a Excel si hay datos
            if len(df_resultado) > 0:
                buf = io.BytesIO()  # Buffer en memoria para el archivo
                df_resultado.to_excel(buf, index=False, sheet_name='Datos_Fuente')
                buf.seek(0)  # Reiniciar posición del buffer para lectura
                
                st.download_button(
                    "📥 Exportar estos datos",
                    data=buf,
                    file_name="datos_fuente_seleccion.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

def generar_alertas(df_filtrado: pd.DataFrame, ind_filtrado: Dict) -> List[Dict]:
    alertas = []
    if ind_filtrado.get('pct_cumplimiento', 100) < UMBRALES_ALERTAS['cumplimiento_minimo']:
        alertas.append({
            'tipo': 'error',
            'titulo': '🔴 Cumplimiento Crítico',
            'mensaje': (
                f"El cumplimiento NNS está en {ind_filtrado.get('pct_cumplimiento', 0)}% "
                f"(meta: {UMBRALES_ALERTAS['cumplimiento_minimo']}%). "
                f"Revisar procesos de producción y logística de inmediato."
            )
        })
    
    if ind_filtrado.get('promedio_desvio_entrega', 0) > UMBRALES_ALERTAS['desvio_entrega_max']:
        alertas.append({
            'tipo': 'warning',
            'titulo': '⚠️ Desvíos Elevados',
            'mensaje': (
                f"El desvío promedio de entrega es de {ind_filtrado.get('promedio_desvio_entrega', 0)} días "
                f"(límite: {UMBRALES_ALERTAS['desvio_entrega_max']} días). "
                f"Evaluar capacidad de transporte y planificación de rutas."
            )
        })
    
    if 'Transportadora' in df_filtrado.columns and len(df_filtrado) > 0:
        perf_transp = df_filtrado.groupby('Transportadora')['Cumple_NNS'].apply(
            lambda x: (x == 'Cumple').sum() / len(x) * 100 if len(x) > 0 else 0
        )
        malas = perf_transp[perf_transp < UMBRALES_ALERTAS['transportadora_min_perf']]
        if len(malas) > 0:
            alertas.append({
                'tipo': 'warning',
                'titulo': '🚚 Transportadoras con Bajo Desempeño',
                'mensaje': (
                    f"{', '.join(malas.index.tolist())} tienen cumplimiento <"
                    f"{UMBRALES_ALERTAS['transportadora_min_perf']}%. "
                    f"Considerar reevaluación de contratos o capacitación."
                )
            })
    
    if 'Cumple_NNS' in df_filtrado.columns and 'PTE' in df_filtrado['Cumple_NNS'].values:
        ptes = df_filtrado[df_filtrado['Cumple_NNS'] == 'PTE']
        if len(ptes) > UMBRALES_ALERTAS['pendientes_max']:
            alertas.append({
                'tipo': 'info',
                'titulo': '⏳ Pendientes Acumulados',
                'mensaje': (
                    f"{len(ptes)} pedidos sin fecha de entrega registrada "
                    f"(umbral: {UMBRALES_ALERTAS['pendientes_max']}). "
                    f"Verificar estado en sistema y actualizar trazabilidad."
                )
            })
    return alertas


def mostrar_alertas(alertas: List[Dict]) -> None:
    if not alertas:
        st.success("✅ No hay alertas activas. Los indicadores están dentro de parámetros normales.")
        return
    css_class = {
        'success': 'rec-success', 
        'warning': 'rec-warning',
        'error': 'rec-error', 
        'info': 'rec-info'
    }
    for alerta in alertas:
        cls = css_class.get(alerta['tipo'], 'rec-info')
        st.markdown(
            f'<div class="rec-card {cls}">'
            f'<strong style="color:#e8ecf4">{alerta["titulo"]}</strong><br>'
            f'<span style="color:#8b9dc3;font-size:0.9rem">{alerta["mensaje"]}</span>'
            f'</div>',
            unsafe_allow_html=True
        )

def mostrar_recomendaciones(processor, df_filtrado: pd.DataFrame) -> None:
    st.markdown("### 💡 Análisis de Mejora")
    recs = processor.get_recomendaciones(df_filtrado)

    if not recs:
        st.info("ℹ️ No hay suficientes datos para generar recomendaciones automatizadas.")
        return

    css_class = {
        'success': 'rec-success', 
        'warning': 'rec-warning',
        'error': 'rec-error', 
        'info': 'rec-info'
    }

    for titulo, cuerpo, tipo in recs:
        cls = css_class.get(tipo, 'rec-info')
        st.markdown(
            f'<div class="rec-card {cls}">'
            f'<strong style="color:#e8ecf4">{titulo}</strong><br>'
            f'<span style="color:#8b9dc3;font-size:0.9rem">{cuerpo}</span>'
            f'</div>',
            unsafe_allow_html=True
        )

def mostrar_tabla_detalle(processor, df_filtrado: pd.DataFrame) -> None:
    st.markdown("### 📋 Detalle de Incumplimientos")
    inc = processor.get_pedidos_incumplimiento(df_filtrado)

    if inc is None or len(inc) == 0:
        st.success("🎉 No hay pedidos con incumplimiento en el período seleccionado.")
        return

    cf1, cf2, cf3 = st.columns(3)
    with cf1:
        ciudades_inc = ['Todas'] + sorted(inc['Ciudad'].dropna().astype(str).unique().tolist())
        c_sel = st.selectbox("📍 Ciudad", ciudades_inc, key='tab_ciudad', index=0)
    with cf2:
        if 'Area_Incumple' in inc.columns:
            areas_inc = ['Todas'] + sorted(inc['Area_Incumple'].dropna().unique().tolist())
            a_sel = st.selectbox("🏢 Área Responsable", areas_inc, key='tab_area', index=0)
        else:
            a_sel = 'Todas'
    with cf3:
        if 'Desvio_Entrega' in inc.columns and len(inc['Desvio_Entrega'].dropna()) > 0:
            min_d = float(inc['Desvio_Entrega'].min())
            max_d = float(inc['Desvio_Entrega'].max())
            d_sel = st.slider("⏱️ Desvío mínimo (días)", min_value=min_d, max_value=max_d,
                              value=min_d, key='tab_desvio')
        else:
            d_sel = 0

    df_t = inc.copy()
    if c_sel != 'Todas':
        df_t = df_t[df_t['Ciudad'].astype(str) == c_sel]
    if a_sel != 'Todas' and 'Area_Incumple' in df_t.columns:
        df_t = df_t[df_t['Area_Incumple'] == a_sel]
    if 'Desvio_Entrega' in df_t.columns:
        df_t = df_t[df_t['Desvio_Entrega'] >= d_sel]

    st.caption(f"Mostrando {len(df_t):,} de {len(inc):,} incumplimientos")

    rename_display = {
        'Fecha': 'Fecha Compra', 'No_Orden': 'No. Orden',
        'Cliente': 'Cliente', 'Producto': 'Producto',
        'Ciudad': 'Ciudad', 'Transportadora': 'Transportadora',
        'No_Guia': 'No. Guía', 'Fecha_Despacho': 'F. Despacho',
        'Fecha_Entrega': 'F. Entrega',
        'Dias_Despacho_Hab': 'Días Despacho', 'Dias_Entrega_Hab': 'Días Entrega',
        'SLA_Entrega': 'SLA', 'Desvio_Despacho': 'Desvío Despacho',
        'Desvio_Entrega': 'Desvío Entrega', 'Area_Incumple': 'Área Responsable',
        'Valor despacho': 'Valor Despacho', 'Causal de Incumplimiento': 'Causal',
        'Categoria': 'Categoría', 'Concepto': 'Tipo'
    }
    df_show = df_t.rename(columns={k: v for k, v in rename_display.items() if k in df_t.columns})

    st.dataframe(df_show, use_container_width=True, hide_index=True)

    col_exp1, col_exp2 = st.columns([1, 4])
    with col_exp1:
        try:
            buf = io.BytesIO()
            df_t.to_excel(buf, index=False, sheet_name='Incumplimientos')
            buf.seek(0)
            st.download_button(
                "📥 Exportar a Excel",
                data=buf,
                file_name="incumplimientos_filtrados.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Descarga los incumplimientos filtrados en formato Excel"
            )
        except Exception as e:
            logger.error(f"Error exportando tabla: {e}")
            st.error("❌ Error al generar archivo de exportación")

def generate_report_advanced(
    df_filtrado: pd.DataFrame, 
    ind_filtrado: Dict, 
    ind_global: Dict,
    processor
) -> io.BytesIO:
    """
    Genera el Mega Reporte Excel con hojas de inteligencia de negocios para toma de decisiones.
    Añadido:
    - Cuellos de botella críticos (peores casos).
    - Matriz de Negociación (Tránsito promedio por Ciudad/Transportadora).
    - Riesgo Financiero (Dinero perdido agrupado por área).
    """
    buf = io.BytesIO()
    
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        # HOJA 1: Resumen Ejecutivo
        resumen = pd.DataFrame({
            'Métrica': [
                'Total Pedidos', 'Cumplimiento NNS', 'Desvío Promedio Entrega', 
                'Valor Total Despachos', 'Transportadoras Activas', 'Ciudades Atendidas'
            ],
            'Valor': [
                ind_filtrado.get('total_pedidos', 0),
                f"{ind_filtrado.get('pct_cumplimiento', 0)}%",
                f"{ind_filtrado.get('promedio_desvio_entrega', 0)} días",
                f"${df_filtrado.get('Valor_num', pd.Series([0])).sum():,.0f}" if 'Valor_num' in df_filtrado.columns else 'N/A',
                df_filtrado['Transportadora'].nunique() if 'Transportadora' in df_filtrado.columns else 0,
                df_filtrado['Ciudad'].nunique() if 'Ciudad' in df_filtrado.columns else 0
            ],
            'Variación vs Global': [
                f"{ind_filtrado.get('total_pedidos', 0) - ind_global.get('total_pedidos', 0):+d}",
                f"{ind_filtrado.get('pct_cumplimiento', 0) - ind_global.get('pct_cumplimiento', 0):+.1f}%",
                '-', '-', '-', '-'
            ]
        })
        resumen.to_excel(writer, sheet_name='📊 Resumen Ejecutivo', index=False)
        
        # HOJA 2: Datos Crudos Filtrados
        df_filtrado.to_excel(writer, sheet_name='📋 Datos Filtrados', index=False)
        
        # HOJA 3: Cuellos de Botella Críticos (Los mayores desvíos de entrega)
        # Esto ayuda al gerente a enfocarse inmediatamente en los peores casos que penalizan la reputación.
        if 'Desvio_Entrega' in df_filtrado.columns:
            # Filtramos aquellos con desvío mayor a 0 y tomamos los 50 peores
            criticos = df_filtrado[df_filtrado['Desvio_Entrega'] > 0].sort_values(by='Desvio_Entrega', ascending=False)
            if len(criticos) > 0:
                criticos.head(50).to_excel(writer, sheet_name='🚨 Cuellos de Botella', index=False)
        
        # HOJA 4: Matriz de Negociación (Transportadora vs Ciudad)
        # Una tabla cruzada (pivot table) que sirve para decidir a qué transportadora darle qué rutas.
        if 'Ciudad' in df_filtrado.columns and 'Transportadora' in df_filtrado.columns and 'Dias_Entrega_Hab' in df_filtrado.columns:
            matriz = pd.pivot_table(
                df_filtrado, 
                values='Dias_Entrega_Hab', 
                index='Ciudad', 
                columns='Transportadora', 
                aggfunc='mean'
            ).round(1).fillna('-')
            matriz.to_excel(writer, sheet_name='🗺️ Matriz Negociación')
            
        # HOJA 5: Riesgo Financiero por Área (Dinero afectado por incumplimiento)
        # Para que el gerente sepa cuánto dinero está en riesgo o ha llegado tarde por cada área.
        if 'Valor_num' in df_filtrado.columns and 'Area_Incumple' in df_filtrado.columns:
            # Agrupar solo los que NO cumplen
            df_retrasos = df_filtrado[df_filtrado['Cumple_NNS'] == 'No cumple']
            if len(df_retrasos) > 0:
                riesgo = df_retrasos.groupby('Area_Incumple').agg(
                    Pedidos_Retrasados=('Cumple_NNS', 'count'),
                    Valor_en_Riesgo=('Valor_num', 'sum')
                ).reset_index()
                riesgo.to_excel(writer, sheet_name='💰 Riesgo Financiero', index=False)

        # HOJA 6: Causales de Incumplimiento
        if 'Causal de Incumplimiento' in df_filtrado.columns:
            df_inc = df_filtrado[df_filtrado['Cumple_NNS'] == 'No cumple']
            if len(df_inc) > 0:
                causal_analysis = df_inc['Causal de Incumplimiento'].value_counts().reset_index()
                causal_analysis.columns = ['Causal', 'Frecuencia']
                causal_analysis.to_excel(writer, sheet_name='🎯 Causales', index=False)
        
        # Dar formato a las cabeceras
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            # No aplicar si es hoja de la Matriz Negociación que tiene multi-índice o índice diferente
            if sheet_name != '🗺️ Matriz Negociación':
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
    
    buf.seek(0)
    return buf
